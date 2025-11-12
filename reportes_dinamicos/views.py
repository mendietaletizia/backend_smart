"""
CU14-CU20: Reportes Dinámicos
"""
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.db.models import Q, Sum, Count, Avg, Max, Min
from django.utils import timezone
from django.conf import settings
from datetime import datetime, timedelta
import json
import logging
import os
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Reporte, ModeloIA, PrediccionVenta
from .interpreter import ReporteInterpreter
from ventas_carrito.models import Venta, DetalleVenta
from productos.models import Producto, Categoria
from autenticacion_usuarios.models import Usuario, Cliente

logger = logging.getLogger(__name__)


@method_decorator(csrf_exempt, name='dispatch')
class SolicitarReporteView(View):
    """
    CU14: Solicitar Reporte por Texto
    CU15: Solicitar Reporte por Voz
    CU16: Interpretar Solicitud
    """
    
    def post(self, request):
        """Procesar solicitud de reporte (texto o voz)"""
        try:
            # Verificar autenticación
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión para generar reportes'
                }, status=401)
            
            user_id = request.session.get('user_id')
            if not user_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Usuario no autenticado'
                }, status=401)
            
            try:
                usuario = Usuario.objects.get(id=user_id)
            except Usuario.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Usuario no encontrado'
                }, status=404)
            
            # Verificar rol del usuario
            es_admin = usuario.id_rol and usuario.id_rol.nombre.lower() == 'administrador'
            
            # Obtener datos
            data = json.loads(request.body)
            texto = data.get('texto', '').strip()
            audio_data = data.get('audio')  # Base64 o bytes (opcional, el frontend ya procesa la voz)
            texto_transcrito = data.get('texto_transcrito')  # Texto transcrito desde el frontend
            
            # CU15: Procesar voz - Si hay texto transcrito, usarlo (viene del reconocimiento de voz del navegador)
            if texto_transcrito:
                texto = texto_transcrito.strip()
                logger.info(f"Reporte solicitado por voz - Usuario: {usuario.id}, Texto: {texto[:50]}...")
            elif audio_data:
                # Si hay audio pero no texto transcrito, intentar procesarlo
                # Por ahora, esperamos que el frontend envíe el texto transcrito
                texto = texto or "Reporte de mis compras"
                logger.warning(f"Audio recibido sin texto transcrito - Usuario: {usuario.id}")
            
            if not texto:
                return JsonResponse({
                    'success': False,
                    'message': 'Texto de solicitud requerido. Por favor, escribe o dicta tu solicitud.'
                }, status=400)
            
            # Obtener filtros enviados desde el frontend (si existen)
            filtros_frontend = data.get('filtros', {})
            
            # CU16: Interpretar solicitud con mejoras de IA
            try:
                interpreter = ReporteInterpreter()
                parametros = interpreter.interpretar(texto)
                
                # Integrar filtros del frontend con los parámetros interpretados
                # Los filtros del frontend tienen prioridad sobre los interpretados del texto
                if filtros_frontend:
                    if 'filtros' not in parametros:
                        parametros['filtros'] = {}
                    # Combinar filtros (los del frontend sobrescriben los interpretados)
                    parametros['filtros'].update(filtros_frontend)
                    # También actualizar fechas si vienen en filtros
                    if 'fecha_desde' in filtros_frontend or 'fecha_hasta' in filtros_frontend:
                        if 'fechas' not in parametros:
                            parametros['fechas'] = {}
                        if 'fecha_desde' in filtros_frontend:
                            parametros['fechas']['desde'] = filtros_frontend['fecha_desde']
                        if 'fecha_hasta' in filtros_frontend:
                            parametros['fechas']['hasta'] = filtros_frontend['fecha_hasta']
                
                # Mejorar detección de "productos que he comprado" vs "mis compras"
                texto_lower = texto.lower()
                es_lista_productos = any(frase in texto_lower for frase in [
                    'productos que he comprado', 'productos comprados', 'productos que compré',
                    'qué he comprado', 'que he comprado', 'qué compré', 'que compré',
                    'lista de productos', 'productos adquiridos', 'artículos comprados'
                ])
                
                if es_lista_productos:
                    parametros['es_lista_productos'] = True
                    parametros['tipo_reporte'] = 'mis_compras'
                    logger.info(f"Detectado: solicitud de lista de productos comprados")
                
                # Validar permisos según el rol
                tipo_reporte = parametros.get('tipo_reporte', 'general')
                
                # Clientes solo pueden solicitar reportes de sus propias compras
                if not es_admin:
                    # Tipos permitidos para clientes
                    tipos_permitidos_cliente = ['mis_compras', 'general']
                    
                    if tipo_reporte not in tipos_permitidos_cliente:
                        # Intentar convertir tipos no permitidos a "mis_compras" de forma inteligente
                        if tipo_reporte == 'financiero':
                            # Si es financiero y es cliente, convertir a mis_compras con enfoque financiero
                            parametros['tipo_reporte'] = 'mis_compras'
                            parametros['enfoque_financiero'] = True
                            parametros['cliente_id'] = usuario.id
                            logger.info(f"Cliente {usuario.id} - Convertido tipo 'financiero' a 'mis_compras' con enfoque financiero")
                        elif tipo_reporte in ['ventas', 'productos', 'clientes', 'inventario']:
                            # Si menciona otros tipos pero es cliente, intentar convertir inteligentemente
                            # Si menciona "productos", puede ser "productos que he comprado"
                            if tipo_reporte == 'productos' and any(palabra in texto_lower for palabra in ['he comprado', 'compré', 'compras', 'mis']):
                                parametros['tipo_reporte'] = 'mis_compras'
                                parametros['es_lista_productos'] = True
                                parametros['cliente_id'] = usuario.id
                                logger.info(f"Cliente {usuario.id} - Convertido 'productos' a 'mis_compras' (lista de productos)")
                            elif tipo_reporte == 'ventas' and any(palabra in texto_lower for palabra in ['mis', 'compras', 'pedidos']):
                                parametros['tipo_reporte'] = 'mis_compras'
                                parametros['cliente_id'] = usuario.id
                                logger.info(f"Cliente {usuario.id} - Convertido 'ventas' a 'mis_compras'")
                            else:
                                # No se puede convertir, denegar
                                return JsonResponse({
                                    'success': False,
                                    'message': f'No tienes permisos para solicitar reportes de tipo "{tipo_reporte}". Solo puedes consultar reportes de tus propias compras. Intenta con: "Mis compras del último mes", "Resumen de mis gastos" o "Productos que he comprado".'
                                }, status=403)
                        elif tipo_reporte == 'general':
                            # Si es "general", intentar inferir que es sobre sus compras
                            # Verificar si el texto menciona compras/pedidos/gastos
                            if any(palabra in texto_lower for palabra in ['compra', 'compras', 'pedido', 'pedidos', 'mis', 'gasto', 'gastos', 'gasté', 'gastado']):
                                parametros['tipo_reporte'] = 'mis_compras'
                                parametros['cliente_id'] = usuario.id
                                # Si menciona gastos, agregar enfoque financiero
                                if any(palabra in texto_lower for palabra in ['gasto', 'gastos', 'gasté', 'gastado', 'cuánto', 'cuanto', 'dinero', 'monto', 'resumen']):
                                    parametros['enfoque_financiero'] = True
                                # Si menciona productos específicamente, marcar como lista de productos
                                if es_lista_productos:
                                    parametros['es_lista_productos'] = True
                                logger.info(f"Cliente {usuario.id} - Convertido 'general' a 'mis_compras' basado en contexto")
                            else:
                                # No se puede inferir, convertir a mis_compras por defecto para clientes
                                parametros['tipo_reporte'] = 'mis_compras'
                                parametros['cliente_id'] = usuario.id
                                logger.info(f"Cliente {usuario.id} - Convertido 'general' a 'mis_compras' por defecto")
                        else:
                            # Tipo no permitido y no convertible
                            return JsonResponse({
                                'success': False,
                                'message': f'No tienes permisos para solicitar reportes de tipo "{tipo_reporte}". Solo puedes consultar reportes de tus propias compras. Intenta con: "Mis compras del último mes", "Resumen de mis gastos" o "Productos que he comprado".'
                            }, status=403)
                    else:
                        # Tipo permitido, asegurar que tenga cliente_id
                        if tipo_reporte == 'mis_compras':
                            parametros['cliente_id'] = usuario.id
            except Exception as e:
                logger.error(f"Error al interpretar solicitud: {str(e)}", exc_info=True)
                return JsonResponse({
                    'success': False,
                    'message': f'Error al interpretar la solicitud: {str(e)}'
                }, status=400)
            
            # Generar reporte
            try:
                generador = GeneradorReporte()
                reporte_data = generador.generar(parametros, usuario)
            except Exception as e:
                logger.error(f"Error al generar reporte: {str(e)}", exc_info=True)
                return JsonResponse({
                    'success': False,
                    'message': f'Error al generar el reporte: {str(e)}'
                }, status=500)
            
            # Convertir fechas a strings para JSON serialization
            parametros_serializados = self._serializar_parametros(parametros)
            
            # Generar nombre descriptivo del reporte basado en la solicitud
            tipo_reporte = parametros.get('tipo_reporte', 'general')
            texto_lower = texto.lower()
            es_lista_productos = parametros.get('es_lista_productos', False)
            enfoque_financiero = parametros.get('enfoque_financiero', False)
            fechas = parametros.get('fechas', {})
            
            # Crear título descriptivo según la solicitud con más variaciones
            if es_lista_productos or any(frase in texto_lower for frase in [
                'productos que he comprado', 'productos comprados', 'productos que compré',
                'qué he comprado', 'que he comprado', 'qué compré', 'que compré',
                'lista de productos', 'productos adquiridos', 'artículos comprados'
            ]):
                # Agregar período si está disponible
                periodo = ""
                if 'desde' in fechas and 'hasta' in fechas:
                    periodo = f" ({fechas.get('desde', '')} - {fechas.get('hasta', '')})"
                nombre_reporte = f"Productos que he comprado{periodo}"
            elif enfoque_financiero or any(frase in texto_lower for frase in [
                'resumen de mis gastos', 'mis gastos', 'cuánto he gastado', 'cuanto he gastado',
                'cuánto gasté', 'cuanto gasté', 'total gastado', 'gastos totales', 'resumen gastos'
            ]):
                periodo = ""
                if 'desde' in fechas and 'hasta' in fechas:
                    periodo = f" ({fechas.get('desde', '')} - {fechas.get('hasta', '')})"
                nombre_reporte = f"Resumen de mis gastos{periodo}"
            elif any(frase in texto_lower for frase in [
                'mis compras', 'historial de compras', 'mis pedidos', 'mis ordenes'
            ]):
                periodo = ""
                if 'desde' in fechas and 'hasta' in fechas:
                    periodo = f" ({fechas.get('desde', '')} - {fechas.get('hasta', '')})"
                nombre_reporte = f"Mis compras{periodo}"
            elif tipo_reporte == 'ventas':
                periodo = ""
                if 'desde' in fechas and 'hasta' in fechas:
                    periodo = f" ({fechas.get('desde', '')} - {fechas.get('hasta', '')})"
                nombre_reporte = f"Reporte de ventas{periodo}"
            elif tipo_reporte == 'productos':
                nombre_reporte = "Reporte de productos"
            elif tipo_reporte == 'clientes':
                nombre_reporte = "Reporte de clientes"
            elif tipo_reporte == 'inventario':
                nombre_reporte = "Reporte de inventario"
            elif tipo_reporte == 'financiero':
                periodo = ""
                if 'desde' in fechas and 'hasta' in fechas:
                    periodo = f" ({fechas.get('desde', '')} - {fechas.get('hasta', '')})"
                nombre_reporte = f"Reporte financiero{periodo}"
            else:
                # Usar el texto original truncado si es descriptivo, o generar título genérico
                if len(texto) < 60 and len(texto) > 5:
                    nombre_reporte = texto.capitalize()
                else:
                    # Intentar extraer palabras clave del texto
                    palabras_clave = []
                    if any(palabra in texto_lower for palabra in ['venta', 'ventas']):
                        palabras_clave.append('ventas')
                    if any(palabra in texto_lower for palabra in ['producto', 'productos']):
                        palabras_clave.append('productos')
                    if any(palabra in texto_lower for palabra in ['cliente', 'clientes']):
                        palabras_clave.append('clientes')
                    
                    if palabras_clave:
                        nombre_reporte = f"Reporte de {', '.join(palabras_clave)}"
                    else:
                        nombre_reporte = f"Reporte generado - {texto[:40]}..." if len(texto) > 40 else f"Reporte: {texto}"
            
            # Guardar filtros aplicados
            filtros_aplicados = {}
            if 'filtros' in parametros:
                filtros_aplicados.update(parametros['filtros'])
            if 'fechas' in parametros:
                filtros_aplicados['fechas'] = parametros['fechas']
            
            # Guardar reporte
            reporte = Reporte.objects.create(
                nombre=nombre_reporte,
                tipo=tipo_reporte,
                descripcion=f"Generado desde: {texto}",
                parametros=parametros_serializados,
                prompt=texto,
                formato=parametros.get('formato', 'pantalla'),
                origen_comando='voz' if (texto_transcrito or audio_data) else 'texto',
                id_usuario=usuario,
                datos=reporte_data,
                filtros_aplicados=filtros_aplicados,
                estado='completado'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Reporte generado exitosamente',
                'reporte': {
                    'id': reporte.id_reporte,
                    'nombre': reporte.nombre,
                    'tipo': reporte.tipo,
                    'formato': reporte.formato,
                    'datos': reporte_data,
                    'fecha': reporte.fecha_generacion.isoformat(),
                    'parametros': parametros
                }
            }, status=201)
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Formato de datos inválido'
            }, status=400)
        except Exception as e:
            logger.error(f"Error en SolicitarReporteView.post: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error interno: {str(e)}'
            }, status=500)
    
    def _serializar_parametros(self, parametros: dict) -> dict:
        """Convierte objetos date/datetime a strings para JSON serialization"""
        import copy
        from datetime import date, datetime
        
        parametros_serializados = copy.deepcopy(parametros)
        
        # Convertir fechas en el diccionario de fechas
        if 'fechas' in parametros_serializados:
            fechas = parametros_serializados['fechas']
            if isinstance(fechas, dict):
                for key in ['desde', 'hasta']:
                    if key in fechas:
                        fecha_val = fechas[key]
                        if isinstance(fecha_val, (date, datetime)):
                            fechas[key] = fecha_val.isoformat()
                        elif not isinstance(fecha_val, str):
                            # Si es otro tipo, intentar convertirlo
                            try:
                                fechas[key] = str(fecha_val)
                            except:
                                if key in fechas:
                                    del fechas[key]
        
        return parametros_serializados


class GeneradorReporte:
    """Generador de reportes dinámicos"""
    
    def generar(self, parametros: dict, usuario: Usuario) -> dict:
        """Genera reporte según parámetros"""
        tipo = parametros.get('tipo_reporte', 'general')
        es_admin = usuario.id_rol and usuario.id_rol.nombre.lower() == 'administrador'
        
        # Reportes específicos para clientes
        if tipo == 'mis_compras':
            return self._generar_reporte_mis_compras(parametros, usuario)
        
        # Reporte general - permitido para todos pero con datos limitados para clientes
        if tipo == 'general':
            if not es_admin:
                # Para clientes, el reporte general se convierte a "mis compras"
                return self._generar_reporte_mis_compras(parametros, usuario)
            else:
                return self._generar_reporte_general(parametros, usuario)
        
        # Reportes solo para administradores (excepto financiero que clientes pueden ver de sus compras)
        if tipo == 'financiero':
            # Los clientes pueden ver reportes financieros de sus propias compras
            if not es_admin:
                # Convertir a reporte de mis compras con enfoque financiero
                parametros['tipo_reporte'] = 'mis_compras'
                parametros['enfoque_financiero'] = True
                return self._generar_reporte_mis_compras(parametros, usuario)
            else:
                return self._generar_reporte_financiero(parametros, usuario)
        
        # Otros reportes solo para administradores
        if not es_admin:
            return {
                'tipo': 'error',
                'datos': [],
                'mensaje': 'No tienes permisos para generar este tipo de reporte. Solo puedes consultar reportes de tus propias compras. Intenta con: "Mis compras del último mes", "Cuánto he gastado" o "Productos que he comprado".'
            }
        
        if tipo == 'ventas':
            return self._generar_reporte_ventas(parametros, usuario)
        elif tipo == 'productos':
            return self._generar_reporte_productos(parametros, usuario)
        elif tipo == 'clientes':
            return self._generar_reporte_clientes(parametros, usuario)
        elif tipo == 'inventario':
            return self._generar_reporte_inventario(parametros, usuario)
        else:
            return self._generar_reporte_general(parametros, usuario)
    
    def _generar_reporte_ventas(self, parametros: dict, usuario: Usuario) -> dict:
        """CU17: Generar reporte de ventas con análisis inteligente"""
        query = Venta.objects.select_related('cliente', 'cliente__id').prefetch_related(
            'detalles', 'detalles__producto', 'detalles__producto__categoria'
        )
        
        # Aplicar filtros inteligentes
        filtros = parametros.get('filtros', {})
        if 'estado' in filtros and filtros['estado']:
            query = query.filter(estado=filtros['estado'])
        if 'metodo_pago' in filtros and filtros['metodo_pago']:
            query = query.filter(metodo_pago=filtros['metodo_pago'])
        if 'categoria' in filtros and filtros['categoria']:
            # Puede ser ID o nombre
            if isinstance(filtros['categoria'], int) or (isinstance(filtros['categoria'], str) and filtros['categoria'].isdigit()):
                categoria_id = int(filtros['categoria'])
                query = query.filter(detalles__producto__categoria__id_categoria=categoria_id).distinct()
            else:
                query = query.filter(detalles__producto__categoria__nombre__icontains=filtros['categoria']).distinct()
        if 'producto' in filtros and filtros['producto']:
            query = query.filter(detalles__producto__nombre__icontains=filtros['producto']).distinct()
        if 'cliente' in filtros and filtros['cliente']:
            # Puede ser ID o nombre
            if isinstance(filtros['cliente'], int) or (isinstance(filtros['cliente'], str) and filtros['cliente'].isdigit()):
                cliente_id = int(filtros['cliente'])
                query = query.filter(cliente__id__id=cliente_id)
            else:
                # Buscar por nombre
                query = query.filter(
                    Q(cliente__id__nombre__icontains=filtros['cliente']) |
                    Q(cliente__id__apellido__icontains=filtros['cliente'])
                )
        if 'monto_minimo' in filtros and filtros['monto_minimo']:
            try:
                monto_min = float(filtros['monto_minimo'])
                query = query.filter(total__gte=monto_min)
            except (ValueError, TypeError):
                logger.warning(f"Error al parsear monto_minimo: {filtros['monto_minimo']}")
        if 'monto_maximo' in filtros and filtros['monto_maximo']:
            try:
                monto_max = float(filtros['monto_maximo'])
                query = query.filter(total__lte=monto_max)
            except (ValueError, TypeError):
                logger.warning(f"Error al parsear monto_maximo: {filtros['monto_maximo']}")
        
        # Aplicar fechas con manejo robusto
        fechas = parametros.get('fechas', {})
        if 'desde' in fechas and fechas['desde']:
            try:
                desde_date = fechas['desde']
                if isinstance(desde_date, str):
                    try:
                        desde_date = datetime.fromisoformat(desde_date.replace('Z', '+00:00'))
                    except:
                        desde_date = datetime.strptime(desde_date, '%Y-%m-%d')
                query = query.filter(fecha_venta__gte=desde_date)
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error al parsear fecha 'desde': {str(e)}")
        if 'hasta' in fechas and fechas['hasta']:
            try:
                hasta_date = fechas['hasta']
                if isinstance(hasta_date, str):
                    try:
                        hasta_date = datetime.fromisoformat(hasta_date.replace('Z', '+00:00'))
                    except:
                        hasta_date = datetime.strptime(hasta_date, '%Y-%m-%d')
                query = query.filter(fecha_venta__lte=hasta_date)
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error al parsear fecha 'hasta': {str(e)}")
        
        # Obtener métricas solicitadas
        metricas = parametros.get('metricas', ['total'])
        agrupacion = parametros.get('agrupacion', [])
        contexto = parametros.get('contexto', {})
        
        # Calcular métricas generales
        total_general = float(query.aggregate(Sum('total'))['total__sum'] or 0)
        cantidad_ventas = query.count()
        promedio_venta = float(query.aggregate(Avg('total'))['total__avg'] or 0)
        max_venta = float(query.aggregate(Max('total'))['total__max'] or 0)
        min_venta = float(query.aggregate(Min('total'))['total__min'] or 0)
        
        # Agrupar según solicitud
        if 'categoria' in agrupacion:
            # Agrupar por categoría con métricas completas
            datos = []
            categorias = Categoria.objects.all()
            for categoria in categorias:
                ventas_cat = query.filter(detalles__producto__categoria=categoria).distinct()
                if ventas_cat.exists():
                    datos.append({
                        'categoria': categoria.nombre,
                        'total_ventas': ventas_cat.count(),
                        'monto_total': float(ventas_cat.aggregate(Sum('total'))['total__sum'] or 0),
                        'promedio_venta': float(ventas_cat.aggregate(Avg('total'))['total__avg'] or 0),
                        'porcentaje_del_total': round((float(ventas_cat.aggregate(Sum('total'))['total__sum'] or 0) / total_general * 100) if total_general > 0 else 0, 2)
                    })
            
            # Ordenar por monto total descendente
            datos.sort(key=lambda x: x['monto_total'], reverse=True)
            
            return {
                'tipo': 'ventas',
                'agrupacion': 'categoria',
                'datos': datos,
                'resumen': {
                    'total_general': total_general,
                    'cantidad_ventas': cantidad_ventas,
                    'promedio_venta': promedio_venta,
                    'max_venta': max_venta,
                    'min_venta': min_venta,
                    'categorias_analizadas': len(datos)
                },
                'metricas_calculadas': metricas
            }
        elif 'dia' in agrupacion or 'semana' in agrupacion or 'mes' in agrupacion:
            # Agrupar por período temporal
            datos = []
            ventas_list = query.order_by('fecha_venta')
            
            # Agrupar por día/semana/mes
            periodo_actual = None
            ventas_periodo = []
            
            for venta in ventas_list:
                fecha = venta.fecha_venta.date()
                if 'dia' in agrupacion:
                    periodo = fecha.isoformat()
                elif 'semana' in agrupacion:
                    # Obtener inicio de semana
                    inicio_semana = fecha - timedelta(days=fecha.weekday())
                    periodo = inicio_semana.isoformat()
                else:  # mes
                    periodo = fecha.replace(day=1).isoformat()
                
                if periodo != periodo_actual:
                    if periodo_actual and ventas_periodo:
                        datos.append({
                            'periodo': periodo_actual,
                            'total_ventas': len(ventas_periodo),
                            'monto_total': sum(float(v.total) for v in ventas_periodo),
                            'promedio_venta': sum(float(v.total) for v in ventas_periodo) / len(ventas_periodo) if ventas_periodo else 0
                        })
                    periodo_actual = periodo
                    ventas_periodo = [venta]
                else:
                    ventas_periodo.append(venta)
            
            # Agregar último período
            if periodo_actual and ventas_periodo:
                datos.append({
                    'periodo': periodo_actual,
                    'total_ventas': len(ventas_periodo),
                    'monto_total': sum(float(v.total) for v in ventas_periodo),
                    'promedio_venta': sum(float(v.total) for v in ventas_periodo) / len(ventas_periodo) if ventas_periodo else 0
                })
            
            return {
                'tipo': 'ventas',
                'agrupacion': agrupacion[0] if agrupacion else 'ninguna',
                'datos': datos,
                'resumen': {
                    'total_general': total_general,
                    'cantidad_ventas': cantidad_ventas,
                    'promedio_venta': promedio_venta,
                    'periodos_analizados': len(datos)
                },
                'metricas_calculadas': metricas
            }
        else:
            # Datos detallados sin agrupar
            ventas = query.order_by('-fecha_venta')[:200]  # Aumentado a 200
            datos = []
            for venta in ventas:
                # Obtener productos de la venta
                productos_venta = []
                for detalle in venta.detalles.all():
                    # Verificar que el producto no sea None
                    if detalle.producto:
                        productos_venta.append({
                            'nombre': detalle.producto.nombre or 'Producto sin nombre',
                            'cantidad': detalle.cantidad,
                            'precio_unitario': float(detalle.precio_unitario),
                            'subtotal': float(detalle.subtotal),
                            'categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else 'Sin categoría'
                        })
                    else:
                        # Si el producto fue eliminado, usar información del detalle
                        productos_venta.append({
                            'nombre': f'Producto eliminado (ID: {detalle.producto_id if hasattr(detalle, "producto_id") else "N/A"})',
                            'cantidad': detalle.cantidad,
                            'precio_unitario': float(detalle.precio_unitario),
                            'subtotal': float(detalle.subtotal),
                            'categoria': 'Sin categoría'
                        })
                
                # Información completa del cliente
                cliente_info = {
                    'id': venta.cliente.id.id,
                    'nombre': f"{venta.cliente.id.nombre} {venta.cliente.id.apellido or ''}".strip(),
                    'email': venta.cliente.id.email,
                    'telefono': venta.cliente.id.telefono or 'No especificado',
                    'direccion': venta.cliente.direccion or 'No especificada',
                    'ciudad': venta.cliente.ciudad or 'No especificada'
                }
                
                datos.append({
                    'id': venta.id_venta,
                    'fecha_formateada': venta.fecha_venta.strftime('%d/%m/%Y %H:%M'),
                    'cliente_nombre': f"{venta.cliente.id.nombre} {venta.cliente.id.apellido or ''}".strip(),
                    'cliente_ciudad': venta.cliente.ciudad or 'No especificada',
                    'total_formateado': f"Bs. {float(venta.total):,.2f}",
                    'estado_display': venta.estado.capitalize(),
                    'metodo_pago_display': venta.metodo_pago.replace('_', ' ').title(),
                    'productos_count': len(productos_venta),
                    # Información adicional en detalles
                    'fecha': venta.fecha_venta.isoformat(),
                    'cliente': cliente_info,
                    'cliente_email': venta.cliente.id.email,
                    'cliente_telefono': venta.cliente.id.telefono or 'No especificado',
                    'cliente_direccion': venta.cliente.direccion or 'No especificada',
                    'total': float(venta.total),
                    'estado': venta.estado,
                    'metodo_pago': venta.metodo_pago,
                    'productos': productos_venta,
                    'direccion_entrega': venta.direccion_entrega or 'No especificada',
                    'notas': venta.notas or ''
                })
            
            return {
                'tipo': 'ventas',
                'datos': datos,
                'resumen': {
                    'total_general': total_general,
                    'cantidad_ventas': cantidad_ventas,
                    'promedio_venta': promedio_venta,
                    'max_venta': max_venta,
                    'min_venta': min_venta,
                    'ventas_mostradas': len(datos),
                    'ventas_totales': cantidad_ventas
                },
                'metricas_calculadas': metricas
            }
    
    def _generar_reporte_productos(self, parametros: dict, usuario: Usuario) -> dict:
        """Generar reporte de productos con información completa"""
        query = Producto.objects.select_related('categoria', 'marca', 'proveedor')
        
        # Aplicar filtros
        filtros = parametros.get('filtros', {})
        if 'categoria' in filtros:
            query = query.filter(categoria__nombre__icontains=filtros['categoria'])
        if 'marca' in filtros:
            query = query.filter(marca__nombre__icontains=filtros['marca'])
        if 'nombre' in filtros:
            query = query.filter(nombre__icontains=filtros['nombre'])
        
        productos = query.order_by('nombre')[:200]  # Aumentado a 200
        
        datos = []
        total_productos = query.count()
        productos_sin_stock = 0
        productos_bajo_stock = 0
        
        for producto in productos:
            from productos.models import Stock
            stock = Stock.objects.filter(producto=producto).first()
            stock_cantidad = stock.cantidad if stock else 0
            
            # Obtener estadísticas de ventas del producto
            ventas_producto = DetalleVenta.objects.filter(producto=producto)
            cantidad_vendida = ventas_producto.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
            monto_total_vendido = float(ventas_producto.aggregate(Sum('subtotal'))['subtotal__sum'] or 0)
            veces_vendido = ventas_producto.values('venta').distinct().count()
            
            if stock_cantidad == 0:
                productos_sin_stock += 1
            elif stock_cantidad < 10:
                productos_bajo_stock += 1
            
            datos.append({
                'nombre': producto.nombre,  # SIEMPRE PRIMERO - Nombre del producto
                'precio': f"Bs. {float(producto.precio):,.2f}",
                'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                'stock': stock_cantidad,
                'marca': producto.marca.nombre if producto.marca else 'Sin marca',
                # Datos adicionales para detalles (no visibles en tabla principal)
                'id': producto.id,
                'descripcion': producto.descripcion or 'Sin descripción',
                'categoria_id': producto.categoria.id_categoria if producto.categoria else None,  # Usar id_categoria en lugar de id
                'precio_numero': float(producto.precio),
                'precio_formateado': f"Bs. {float(producto.precio):,.2f}",
                'stock_disponible': stock_cantidad > 0,
                'stock_bajo': stock_cantidad < 10 and stock_cantidad > 0,
                'marca_id': producto.marca.id_marca if producto.marca else None,  # Usar id_marca en lugar de id
                'proveedor': producto.proveedor.nombre if producto.proveedor else 'Sin proveedor',
                'proveedor_id': producto.proveedor.id_proveedor if producto.proveedor else None,  # Usar id_proveedor en lugar de id
                'cantidad_vendida': cantidad_vendida,
                'monto_total_vendido': round(monto_total_vendido, 2),
                'veces_vendido': veces_vendido,
                'fecha_actualizacion_stock': stock.fecha_actualizacion.isoformat() if stock else None
            })
        
        # Ordenar según solicitud
        agrupacion = parametros.get('agrupacion', [])
        
        # Si se solicita agrupación por categoría, agrupar realmente
        if 'categoria' in agrupacion:
            # Agrupar por categoría manteniendo todos los productos
            datos.sort(key=lambda x: (x['categoria'] or 'Sin categoría', x['nombre']))
        elif 'precio' in agrupacion or 'monto' in agrupacion:
            # Ordenar por precio_numero (número) no por precio (string formateado)
            datos.sort(key=lambda x: x.get('precio_numero', 0), reverse=True)
        elif 'stock' in agrupacion:
            datos.sort(key=lambda x: x['stock'])
        elif 'ventas' in agrupacion or 'popularidad' in agrupacion or 'más vendidos' in str(parametros.get('tipo_reporte', '')).lower() or 'mas vendidos' in str(parametros.get('tipo_reporte', '')).lower():
            # Ordenar por cantidad vendida (productos más vendidos)
            datos.sort(key=lambda x: x.get('cantidad_vendida', 0), reverse=True)
        else:
            # Por defecto, ordenar por nombre
            datos.sort(key=lambda x: x['nombre'])
        
        return {
            'tipo': 'productos',
            'datos': datos,
            'resumen': {
                'total_productos': total_productos,
                'productos_sin_stock': productos_sin_stock,
                'productos_bajo_stock': productos_bajo_stock,
                'productos_disponibles': total_productos - productos_sin_stock,
                'productos_mostrados': len(datos)
            },
            'metricas_calculadas': parametros.get('metricas', ['total', 'stock', 'ventas'])
        }
    
    def _generar_reporte_clientes(self, parametros: dict, usuario: Usuario) -> dict:
        """Generar reporte de clientes con información completa"""
        query = Cliente.objects.select_related('id')
        
        # Aplicar filtros si existen
        filtros = parametros.get('filtros', {})
        if 'ciudad' in filtros:
            query = query.filter(ciudad__icontains=filtros['ciudad'])
        if 'nombre' in filtros:
            query = query.filter(id__nombre__icontains=filtros['nombre'])
        
        # Aplicar fechas si se solicitan (fecha de registro o última compra)
        fechas = parametros.get('fechas', {})
        
        clientes = query.order_by('id__nombre')[:200]  # Aumentado a 200
        
        datos = []
        total_clientes = query.count()
        total_ventas_general = 0
        total_monto_general = 0.0
        
        for cliente in clientes:
            # Obtener todas las ventas del cliente
            ventas_cliente = Venta.objects.filter(cliente=cliente)
            
            # Aplicar filtro de fecha si existe (para última compra)
            if 'desde' in fechas and fechas['desde']:
                try:
                    desde_date = fechas['desde']
                    if isinstance(desde_date, str):
                        try:
                            desde_date = datetime.fromisoformat(desde_date.replace('Z', '+00:00'))
                        except:
                            desde_date = datetime.strptime(desde_date, '%Y-%m-%d')
                    ventas_cliente = ventas_cliente.filter(fecha_venta__gte=desde_date)
                except (ValueError, TypeError, AttributeError) as e:
                    logger.warning(f"Error al parsear fecha 'desde' en reporte clientes: {str(e)}")
            
            if 'hasta' in fechas and fechas['hasta']:
                try:
                    hasta_date = fechas['hasta']
                    if isinstance(hasta_date, str):
                        try:
                            hasta_date = datetime.fromisoformat(hasta_date.replace('Z', '+00:00'))
                        except:
                            hasta_date = datetime.strptime(hasta_date, '%Y-%m-%d')
                    ventas_cliente = ventas_cliente.filter(fecha_venta__lte=hasta_date)
                except (ValueError, TypeError, AttributeError) as e:
                    logger.warning(f"Error al parsear fecha 'hasta' en reporte clientes: {str(e)}")
            
            ventas_count = ventas_cliente.count()
            total_compras = float(ventas_cliente.aggregate(Sum('total'))['total__sum'] or 0)
            promedio_compra = float(ventas_cliente.aggregate(Avg('total'))['total__avg'] or 0) if ventas_count > 0 else 0
            max_compra = float(ventas_cliente.aggregate(Max('total'))['total__max'] or 0)
            min_compra = float(ventas_cliente.aggregate(Min('total'))['total__min'] or 0)
            
            # Obtener última compra
            ultima_venta = ventas_cliente.order_by('-fecha_venta').first()
            ultima_compra_fecha = ultima_venta.fecha_venta.isoformat() if ultima_venta else None
            ultima_compra_monto = float(ultima_venta.total) if ultima_venta else 0
            
            # Obtener primera compra (fecha de registro efectivo como cliente activo)
            primera_venta = ventas_cliente.order_by('fecha_venta').first()
            primera_compra_fecha = primera_venta.fecha_venta.isoformat() if primera_venta else None
            
            # Obtener productos más comprados por este cliente
            productos_mas_comprados = []
            if ventas_cliente.exists():
                from django.db.models import Sum as SumModel
                productos_stats = DetalleVenta.objects.filter(
                    venta__cliente=cliente
                ).values('producto__nombre', 'producto__categoria__nombre').annotate(
                    total_cantidad=SumModel('cantidad'),
                    total_monto=SumModel('subtotal')
                ).order_by('-total_cantidad')[:5]
                
                productos_mas_comprados = [
                    {
                        'producto': item['producto__nombre'],
                        'categoria': item['producto__categoria__nombre'] or 'Sin categoría',
                        'cantidad_total': item['total_cantidad'],
                        'monto_total': float(item['total_monto'])
                    }
                    for item in productos_stats
                ]
            
            total_ventas_general += ventas_count
            total_monto_general += total_compras
            
            # Formatear última compra para mostrar
            ultima_compra_display = None
            if ultima_venta:
                ultima_compra_display = ultima_venta.fecha_venta.strftime('%d/%m/%Y')
            
            # Solo incluir información básica y esencial
            datos.append({
                'nombre': f"{cliente.id.nombre} {cliente.id.apellido or ''}".strip(),  # SIEMPRE PRIMERO
                'email': cliente.id.email or 'No especificado',
                'telefono': cliente.id.telefono or 'No especificado',
                'direccion': cliente.direccion or 'No especificada',
                'ciudad': cliente.ciudad or 'No especificada',
                'total_compras': f"Bs. {round(total_compras, 2):,.2f}",
                'ultima_compra': ultima_compra_display or 'Sin compras',
                # Datos adicionales ocultos (solo para detalles si se necesitan)
                'id': cliente.id.id,
                'estado': 'Activo' if cliente.id.estado else 'Inactivo',
                'ventas_count': ventas_count,
                'total_compras_numero': round(total_compras, 2),
                'promedio_compra': round(promedio_compra, 2),
                'max_compra': round(max_compra, 2),
                'min_compra': round(min_compra, 2),
                'ultima_compra_fecha': ultima_compra_fecha,
                'ultima_compra_monto': round(ultima_compra_monto, 2),
                'primera_compra_fecha': primera_compra_fecha,
                'productos_mas_comprados': productos_mas_comprados,
                'es_cliente_frecuente': ventas_count >= 5,
                'es_cliente_vip': total_compras >= 1000
            })
        
        # Ordenar según solicitud
        agrupacion = parametros.get('agrupacion', [])
        if 'total_compras' in agrupacion or 'monto' in agrupacion:
            datos.sort(key=lambda x: x.get('total_compras_numero', 0), reverse=True)
        elif 'ventas_count' in agrupacion or 'frecuencia' in agrupacion:
            datos.sort(key=lambda x: x.get('ventas_count', 0), reverse=True)
        elif 'ciudad' in agrupacion:
            datos.sort(key=lambda x: x.get('ciudad', ''))
        else:
            # Por defecto, ordenar por nombre
            datos.sort(key=lambda x: x.get('nombre', ''))
        
        return {
            'tipo': 'clientes',
            'datos': datos,
            'resumen': {
                'total_clientes': total_clientes,
                'total_ventas': total_ventas_general,
                'total_monto_general': round(total_monto_general, 2),
                'promedio_ventas_por_cliente': round(total_ventas_general / total_clientes, 2) if total_clientes > 0 else 0,
                'promedio_monto_por_cliente': round(total_monto_general / total_clientes, 2) if total_clientes > 0 else 0,
                'clientes_frecuentes': len([c for c in datos if c['es_cliente_frecuente']]),
                'clientes_vip': len([c for c in datos if c['es_cliente_vip']]),
                'clientes_mostrados': len(datos)
            },
            'metricas_calculadas': parametros.get('metricas', ['total', 'promedio', 'cantidad'])
        }
    
    def _generar_reporte_inventario(self, parametros: dict, usuario: Usuario) -> dict:
        """Generar reporte de inventario"""
        from productos.models import Stock
        
        query = Stock.objects.select_related('producto', 'producto__categoria')
        
        # Filtrar solo stocks con productos válidos y ordenar
        stocks = query.filter(producto__isnull=False).order_by('producto__nombre')[:100]
        
        datos = []
        for stock in stocks:
            # Verificar que el producto no sea None
            if stock.producto:
                datos.append({
                    'producto_id': stock.producto.id,
                    'producto_nombre': stock.producto.nombre or 'Producto sin nombre',
                    'categoria': stock.producto.categoria.nombre if stock.producto.categoria else 'Sin categoría',
                    'cantidad': stock.cantidad,
                    'fecha_actualizacion': stock.fecha_actualizacion.isoformat()
                })
            else:
                # Si el producto fue eliminado
                datos.append({
                    'producto_id': None,
                    'producto_nombre': f'Producto eliminado (Stock ID: {stock.id})',
                    'categoria': 'Sin categoría',
                    'cantidad': stock.cantidad,
                    'fecha_actualizacion': stock.fecha_actualizacion.isoformat()
                })
        
        return {
            'tipo': 'inventario',
            'datos': datos,
            'total_productos': query.count()
        }
    
    def _generar_reporte_financiero(self, parametros: dict, usuario: Usuario) -> dict:
        """Generar reporte financiero con estructura completa para el frontend"""
        query = Venta.objects.filter(estado='completada')
        
        # Aplicar fechas con manejo robusto
        fechas = parametros.get('fechas', {})
        if 'desde' in fechas and fechas['desde']:
            try:
                desde_date = fechas['desde']
                if isinstance(desde_date, str):
                    try:
                        desde_date = datetime.fromisoformat(desde_date.replace('Z', '+00:00'))
                    except:
                        desde_date = datetime.strptime(desde_date, '%Y-%m-%d')
                query = query.filter(fecha_venta__gte=desde_date)
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error al parsear fecha 'desde' en reporte financiero: {str(e)}")
        if 'hasta' in fechas and fechas['hasta']:
            try:
                hasta_date = fechas['hasta']
                if isinstance(hasta_date, str):
                    try:
                        hasta_date = datetime.fromisoformat(hasta_date.replace('Z', '+00:00'))
                    except:
                        hasta_date = datetime.strptime(hasta_date, '%Y-%m-%d')
                query = query.filter(fecha_venta__lte=hasta_date)
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error al parsear fecha 'hasta' en reporte financiero: {str(e)}")
        
        total_ingresos = float(query.aggregate(Sum('total'))['total__sum'] or 0)
        cantidad_ventas = query.count()
        promedio_venta = float(query.aggregate(Avg('total'))['total__avg'] or 0)
        max_venta = float(query.aggregate(Max('total'))['total__max'] or 0)
        min_venta = float(query.aggregate(Min('total'))['total__min'] or 0)
        
        # Por método de pago - Solo Stripe
        por_metodo = []
        metodos_pago_map = {
            'stripe': 'Pago Online (Stripe)'
        }
        
        for metodo_key, metodo_display in metodos_pago_map.items():
            ventas_metodo = query.filter(metodo_pago=metodo_key)
            cantidad = ventas_metodo.count()
            total_metodo = float(ventas_metodo.aggregate(Sum('total'))['total__sum'] or 0)
            if cantidad > 0 or total_metodo > 0:  # Solo incluir si hay datos
                porcentaje = round((total_metodo / total_ingresos * 100) if total_ingresos > 0 else 0, 2)
                por_metodo.append({
                    'metodo_pago': metodo_display,
                    'metodo_pago_key': metodo_key,
                    'cantidad_ventas': cantidad,
                    'total': total_metodo,
                    'total_formateado': f"Bs. {total_metodo:,.2f}",
                    'porcentaje': porcentaje
                })
        
        # Ordenar por total descendente
        por_metodo.sort(key=lambda x: x['total'], reverse=True)
        
        # Crear estructura de datos para el frontend
        datos = []
        if por_metodo:
            datos = por_metodo
        
        return {
            'tipo': 'financiero',
            'datos': datos,  # Array para que el frontend pueda renderizar
            'resumen': {
                'total_ingresos': total_ingresos,
                'total_ingresos_formateado': f"Bs. {total_ingresos:,.2f}",
                'cantidad_ventas': cantidad_ventas,
                'promedio_venta': promedio_venta,
                'promedio_venta_formateado': f"Bs. {promedio_venta:,.2f}",
                'max_venta': max_venta,
                'max_venta_formateado': f"Bs. {max_venta:,.2f}",
                'min_venta': min_venta,
                'min_venta_formateado': f"Bs. {min_venta:,.2f}",
                'por_metodo_pago': por_metodo  # Mantener para compatibilidad
            }
        }
    
    def _generar_reporte_mis_compras(self, parametros: dict, usuario: Usuario) -> dict:
        """Generar reporte de compras del cliente con datos completos"""
        try:
            # Intentar obtener el cliente usando el id del usuario
            try:
                cliente = Cliente.objects.get(id=usuario)
            except Cliente.DoesNotExist:
                # Si no existe como Cliente, intentar buscar por id_usuario
                try:
                    cliente = Cliente.objects.get(id_usuario=usuario)
                except Cliente.DoesNotExist:
                    logger.warning(f"Cliente no encontrado para usuario {usuario.id}")
                    return {
                        'tipo': 'mis_compras',
                        'datos': [],
                        'resumen': {
                            'total_general': 0,
                            'total_general_formateado': '$0.00',
                            'cantidad_compras': 0,
                            'promedio_compra': 0,
                            'promedio_compra_formateado': '$0.00',
                            'mensaje': 'No se encontró información de cliente. Asegúrate de tener compras registradas.'
                        }
                    }
        except Exception as e:
            logger.error(f"Error al obtener cliente para usuario {usuario.id}: {str(e)}", exc_info=True)
            return {
                'tipo': 'mis_compras',
                'datos': [],
                'resumen': {
                    'total_general': 0,
                    'total_general_formateado': '$0.00',
                    'cantidad_compras': 0,
                    'promedio_compra': 0,
                    'promedio_compra_formateado': '$0.00',
                    'mensaje': f'Error al obtener información: {str(e)}'
                }
            }
        
        query = Venta.objects.filter(cliente=cliente).select_related('cliente', 'cliente__id').prefetch_related(
            'detalles', 'detalles__producto', 'detalles__producto__categoria'
        )
        
        # Aplicar filtros de fecha
        fechas = parametros.get('fechas', {})
        if 'desde' in fechas and fechas['desde']:
            try:
                desde_date = fechas['desde']
                if isinstance(desde_date, str):
                    try:
                        desde_date = datetime.fromisoformat(desde_date.replace('Z', '+00:00'))
                    except:
                        desde_date = datetime.strptime(desde_date, '%Y-%m-%d')
                if isinstance(desde_date, datetime):
                    desde_date = desde_date.date()
                query = query.filter(fecha_venta__date__gte=desde_date)
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error al parsear fecha 'desde': {str(e)}")
        if 'hasta' in fechas and fechas['hasta']:
            try:
                hasta_date = fechas['hasta']
                if isinstance(hasta_date, str):
                    try:
                        hasta_date = datetime.fromisoformat(hasta_date.replace('Z', '+00:00'))
                    except:
                        hasta_date = datetime.strptime(hasta_date, '%Y-%m-%d')
                if isinstance(hasta_date, datetime):
                    hasta_date = hasta_date.date()
                query = query.filter(fecha_venta__date__lte=hasta_date)
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error al parsear fecha 'hasta': {str(e)}")
        
        # Aplicar filtros de estado
        filtros = parametros.get('filtros', {})
        if 'estado' in filtros:
            query = query.filter(estado=filtros['estado'])
        if 'metodo_pago' in filtros:
            query = query.filter(metodo_pago=filtros['metodo_pago'])
        if 'categoria' in filtros:
            query = query.filter(detalles__producto__categoria__nombre__icontains=filtros['categoria']).distinct()
        
        # Calcular métricas generales
        total_general = float(query.aggregate(Sum('total'))['total__sum'] or 0)
        cantidad_compras = query.count()
        promedio_compra = float(query.aggregate(Avg('total'))['total__avg'] or 0) if cantidad_compras > 0 else 0
        max_compra = float(query.aggregate(Max('total'))['total__max'] or 0)
        min_compra = float(query.aggregate(Min('total'))['total__min'] or 0)
        
        # Agrupar por categoría si se solicita
        agrupacion = parametros.get('agrupacion', [])
        metricas = parametros.get('metricas', ['total'])
        
        if 'categoria' in agrupacion:
            datos = []
            categorias = Categoria.objects.all()
            for categoria in categorias:
                ventas_cat = query.filter(detalles__producto__categoria=categoria).distinct()
                if ventas_cat.exists():
                    monto_cat = float(ventas_cat.aggregate(Sum('total'))['total__sum'] or 0)
                    datos.append({
                        'categoria': categoria.nombre,
                        'total_compras': ventas_cat.count(),
                        'monto_total': monto_cat,
                        'promedio_compra': float(ventas_cat.aggregate(Avg('total'))['total__avg'] or 0),
                        'porcentaje_del_total': round((monto_cat / total_general * 100) if total_general > 0 else 0, 2)
                    })
            
            # Ordenar por monto total descendente
            datos.sort(key=lambda x: x['monto_total'], reverse=True)
            
            return {
                'tipo': 'mis_compras',
                'agrupacion': 'categoria',
                'datos': datos,
                'resumen': {
                    'total_general': total_general,
                    'cantidad_compras': cantidad_compras,
                    'promedio_compra': promedio_compra,
                    'max_compra': max_compra,
                    'min_compra': min_compra,
                    'categorias_analizadas': len(datos)
                },
                'metricas_calculadas': metricas
            }
        else:
            # Verificar si se solicita lista de productos (no ventas)
            # Usar el parámetro es_lista_productos si está disponible, sino verificar texto_original
            es_lista_productos = parametros.get('es_lista_productos', False)
            if not es_lista_productos:
                texto_original = parametros.get('texto_original', '').lower()
                es_lista_productos = any(palabra in texto_original for palabra in [
                    'productos que he comprado', 'productos comprados', 'qué he comprado', 
                    'que he comprado', 'productos que compré', 'lista de productos comprados',
                    'productos adquiridos', 'artículos comprados'
                ])
            
            if es_lista_productos:
                # Generar lista de productos comprados (no ventas)
                ventas = query.order_by('-fecha_venta')[:200]
                productos_comprados = {}  # Usar dict para agrupar por producto
                
                for venta in ventas:
                    for detalle in venta.detalles.all():
                        if detalle.producto:
                            producto_id = detalle.producto.id
                            if producto_id not in productos_comprados:
                                productos_comprados[producto_id] = {
                                    'nombre': detalle.producto.nombre or 'Producto sin nombre',
                                    'categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else 'Sin categoría',
                                    'cantidad_total': 0,
                                    'monto_total': 0.0,
                                    'precio_unitario_promedio': 0.0,
                                    'primera_compra': venta.fecha_venta,
                                    'ultima_compra': venta.fecha_venta,
                                    'veces_comprado': 0
                                }
                            
                            prod = productos_comprados[producto_id]
                            prod['cantidad_total'] += detalle.cantidad
                            prod['monto_total'] += float(detalle.subtotal)
                            prod['veces_comprado'] += 1
                            if venta.fecha_venta < prod['primera_compra']:
                                prod['primera_compra'] = venta.fecha_venta
                            if venta.fecha_venta > prod['ultima_compra']:
                                prod['ultima_compra'] = venta.fecha_venta
                
                # Convertir a lista y calcular promedios
                datos = []
                for prod_id, prod_data in productos_comprados.items():
                    prod_data['precio_unitario_promedio'] = prod_data['monto_total'] / prod_data['cantidad_total'] if prod_data['cantidad_total'] > 0 else 0
                    datos.append({
                        'nombre': prod_data['nombre'],  # SIEMPRE PRIMERO - Nombre del producto
                        'precio_total': f"Bs. {prod_data['monto_total']:,.2f}",
                        'precio_unitario': f"Bs. {prod_data['precio_unitario_promedio']:,.2f}",
                        'cantidad': prod_data['cantidad_total'],
                        'fecha': prod_data['ultima_compra'].strftime('%d/%m/%Y'),
                        'categoria': prod_data['categoria'],
                        # Datos adicionales para detalles (no visibles en tabla principal)
                        'monto_total': prod_data['monto_total'],
                        'precio_unitario_promedio': prod_data['precio_unitario_promedio'],
                        'fecha_primera_compra': prod_data['primera_compra'].strftime('%d/%m/%Y'),
                        'veces_comprado': prod_data['veces_comprado']
                    })
                
                # Ordenar por fecha de última compra (más reciente primero)
                datos.sort(key=lambda x: x['fecha'], reverse=True)
            else:
                # Datos detallados de ventas sin agrupar
                ventas = query.order_by('-fecha_venta')[:200]
                datos = []
                for venta in ventas:
                    productos_list = []
                    nombres_productos = []  # Lista de nombres para mostrar en la tabla principal
                    for detalle in venta.detalles.all():
                        # Verificar que el producto no sea None
                        if detalle.producto:
                            nombre_producto = detalle.producto.nombre or 'Producto sin nombre'
                            productos_list.append({
                                'nombre': nombre_producto,
                                'cantidad': detalle.cantidad,
                                'precio_unitario': float(detalle.precio_unitario),
                                'subtotal': float(detalle.subtotal),
                                'categoria': detalle.producto.categoria.nombre if detalle.producto.categoria else 'Sin categoría'
                            })
                            # Agregar nombre a la lista (con cantidad si hay más de uno)
                            if detalle.cantidad > 1:
                                nombres_productos.append(f"{nombre_producto} (x{detalle.cantidad})")
                            else:
                                nombres_productos.append(nombre_producto)
                        else:
                            # Si el producto fue eliminado, usar información del detalle
                            nombre_eliminado = f'Producto eliminado (ID: {detalle.producto_id if hasattr(detalle, "producto_id") else "N/A"})'
                            productos_list.append({
                                'nombre': nombre_eliminado,
                                'cantidad': detalle.cantidad,
                                'precio_unitario': float(detalle.precio_unitario),
                                'subtotal': float(detalle.subtotal),
                                'categoria': 'Sin categoría'
                            })
                            if detalle.cantidad > 1:
                                nombres_productos.append(f"{nombre_eliminado} (x{detalle.cantidad})")
                            else:
                                nombres_productos.append(nombre_eliminado)
                    
                    # Crear string con todos los nombres de productos (para mostrar en tabla principal)
                    nombres_display = ', '.join(nombres_productos)
                    # Si hay muchos productos, truncar pero mostrar los primeros
                    if len(nombres_display) > 100:
                        nombres_display = nombres_display[:97] + '...'
                    
                    datos.append({
                        'nombre': nombres_display,  # NOMBRE SIEMPRE PRIMERO - Nombres de productos
                        'fecha': venta.fecha_venta.strftime('%d/%m/%Y %H:%M'),
                        'total': f"Bs. {float(venta.total):,.2f}",
                        'estado': venta.estado,
                        'estado_display': venta.estado.capitalize(),
                        # Información adicional en detalles (no visible en tabla principal)
                        'fecha_iso': venta.fecha_venta.isoformat(),
                        'total_numero': float(venta.total),
                        'metodo_pago': venta.metodo_pago,
                        'metodo_pago_display': venta.metodo_pago.replace('_', ' ').title(),
                        'productos': productos_list,
                        'productos_count': len(productos_list),
                        'total_productos_cantidad': sum(p['cantidad'] for p in productos_list),
                        'direccion_entrega': venta.direccion_entrega or 'No especificada',
                        'notas': venta.notas or '',
                        'categorias_incluidas': list(set([p['categoria'] for p in productos_list])),
                        'nombres_productos': nombres_productos  # Lista completa de nombres
                    })
            
            # Si tiene enfoque financiero, agregar más métricas financieras
            resumen = {
                'total_general': total_general,
                'total_general_formateado': f"Bs. {total_general:,.2f}",
                'cantidad_compras': cantidad_compras,
                'promedio_compra': promedio_compra,
                'promedio_compra_formateado': f"Bs. {promedio_compra:,.2f}",
                'max_compra': max_compra,
                'max_compra_formateado': f"Bs. {max_compra:,.2f}",
                'min_compra': min_compra,
                'min_compra_formateado': f"Bs. {min_compra:,.2f}",
                'compras_mostradas': len(datos),
                'compras_totales': cantidad_compras
            }
            
            # Agregar métricas financieras adicionales si se solicita
            if parametros.get('enfoque_financiero', False):
                # Agrupar por método de pago - Solo Stripe
                por_metodo = {}
                for metodo in ['stripe']:
                    ventas_metodo = query.filter(metodo_pago=metodo)
                    if ventas_metodo.exists():
                        por_metodo[metodo] = {
                            'cantidad': ventas_metodo.count(),
                            'total': float(ventas_metodo.aggregate(Sum('total'))['total__sum'] or 0),
                            'total_formateado': f"${float(ventas_metodo.aggregate(Sum('total'))['total__sum'] or 0):,.2f}"
                        }
                
                # Agrupar por estado
                por_estado = {}
                for estado in ['completada', 'pendiente', 'cancelada']:
                    ventas_estado = query.filter(estado=estado)
                    if ventas_estado.exists():
                        por_estado[estado] = {
                            'cantidad': ventas_estado.count(),
                            'total': float(ventas_estado.aggregate(Sum('total'))['total__sum'] or 0),
                            'total_formateado': f"${float(ventas_estado.aggregate(Sum('total'))['total__sum'] or 0):,.2f}"
                        }
                
                resumen['por_metodo_pago'] = por_metodo
                resumen['por_estado'] = por_estado
                resumen['tipo_reporte'] = 'financiero_personal'
            
            return {
                'tipo': 'mis_compras',
                'datos': datos,
                'resumen': resumen,
                'metricas_calculadas': metricas,
                'enfoque_financiero': parametros.get('enfoque_financiero', False)
            }
    
    def _generar_reporte_general(self, parametros: dict, usuario: Usuario) -> dict:
        """Generar reporte general"""
        return {
            'tipo': 'general',
            'mensaje': 'Reporte general',
            'parametros': parametros
        }


@method_decorator(csrf_exempt, name='dispatch')
class DescargarReporteView(View):
    """
    CU18: Generar Reporte en PDF o Excel
    CU20: Descargar o Visualizar Reporte
    """
    
    def get(self, request, reporte_id):
        """Descargar reporte en formato PDF o Excel"""
        try:
            formato = request.GET.get('formato', 'pdf')
            
            try:
                reporte = Reporte.objects.get(id_reporte=reporte_id)
            except Reporte.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Reporte no encontrado'
                }, status=404)
            
            if formato == 'pdf':
                return self._generar_pdf(reporte)
            elif formato == 'excel':
                return self._generar_excel(reporte)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Formato no soportado. Use pdf o excel'
                }, status=400)
                
        except Exception as e:
            logger.error(f"Error en DescargarReporteView.get: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error interno: {str(e)}'
            }, status=500)
    
    def _generar_pdf(self, reporte: Reporte):
        """Generar PDF del reporte con formato mejorado"""
        from reportlab.lib.units import inch
        from reportlab.lib.styles import ParagraphStyle
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, 
                               leftMargin=0.5*inch, rightMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#0066FF'),
            spaceAfter=12,
            alignment=0  # Left align
        )
        
        # Título
        story.append(Paragraph(f"<b>{reporte.nombre}</b>", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Información del reporte (solo información relevante)
        # Convertir a zona horaria local si está en UTC
        from django.utils import timezone
        fecha_generacion = reporte.fecha_generacion
        if timezone.is_aware(fecha_generacion):
            fecha_generacion = timezone.localtime(fecha_generacion)
        fecha_formateada = fecha_generacion.strftime('%d/%m/%Y %H:%M:%S')
        
        info_data = [
            ['Fecha de Generación:', fecha_formateada],
            ['Origen:', reporte.get_origen_comando_display()],
        ]
        
        if reporte.prompt:
            info_data.append(['Solicitud Original:', reporte.prompt[:100] + ('...' if len(reporte.prompt) > 100 else '')])
        
        info_table = Table(info_data, colWidths=[2.5*inch, 4.5*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.4*inch))
        
        # Datos del reporte
        datos = reporte.datos
        if isinstance(datos, dict):
            # Mostrar resumen si existe
            if 'resumen' in datos and datos['resumen']:
                story.append(Paragraph("<b>📊 Resumen</b>", styles['Heading2']))
                story.append(Spacer(1, 0.15*inch))
                
                resumen = datos['resumen']
                resumen_data = []
                # Campos a excluir del resumen
                campos_excluir = [
                    'compras_mostradas', 'compras_totales', 'categorias_analizadas', 
                    'clientes_mostrados', 'mensaje', 'compras_totales'
                ]
                
                # Campos prioritarios a mostrar primero
                campos_prioritarios = [
                    'total_general', 'total_general_formateado',
                    'cantidad_compras', 'promedio_compra', 'promedio_compra_formateado',
                    'max_compra', 'max_compra_formateado',
                    'min_compra', 'min_compra_formateado'
                ]
                
                # Primero agregar campos prioritarios formateados
                for key in campos_prioritarios:
                    if key in resumen and key not in campos_excluir:
                        value = resumen[key]
                        if isinstance(value, (dict, list)):
                            continue
                        
                        # Si es formateado, usarlo directamente
                        if key.endswith('_formateado'):
                            label = key.replace('_formateado', '').replace('_', ' ').title().strip()
                            resumen_data.append([label, str(value)])
                        elif isinstance(value, (int, float)):
                            # Formatear si no hay versión formateada
                            key_formateado = key + '_formateado'
                            if key_formateado not in resumen:
                                if 'total' in key.lower() or 'monto' in key.lower() or 'precio' in key.lower() or 'compra' in key.lower():
                                    value_str = f"Bs. {value:,.2f}"
                                else:
                                    value_str = f"{value:,}"
                                label = key.replace('_', ' ').title().strip()
                                resumen_data.append([label, value_str])
                
                # Luego agregar otros campos relevantes (no prioritarios, no excluidos)
                for key, value in resumen.items():
                    if key in campos_excluir or key in campos_prioritarios:
                        continue
                    if isinstance(value, (dict, list)):
                        continue
                    
                    # Si hay versión formateada, usar esa
                    key_formateado = key + '_formateado'
                    if key_formateado in resumen:
                        label = key.replace('_', ' ').title().strip()
                        resumen_data.append([label, str(resumen[key_formateado])])
                    elif not key.endswith('_formateado') and not key.endswith('_display'):
                        # Solo agregar si no es un campo formateado y no tiene versión formateada
                        if isinstance(value, (int, float)):
                            if 'total' in key.lower() or 'monto' in key.lower() or 'precio' in key.lower() or 'compra' in key.lower():
                                value_str = f"Bs. {value:,.2f}"
                            else:
                                value_str = f"{value:,}"
                        else:
                            value_str = str(value)
                        label = key.replace('_', ' ').title().strip()
                        resumen_data.append([label, value_str])
                
                if resumen_data:
                    resumen_table = Table(resumen_data, colWidths=[3*inch, 4*inch])
                    resumen_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066FF')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ]))
                    story.append(resumen_table)
                    story.append(Spacer(1, 0.3*inch))
            
            # Mostrar datos principales
            if 'datos' in datos and datos['datos']:
                story.append(Paragraph("<b>📋 Detalles</b>", styles['Heading2']))
                story.append(Spacer(1, 0.15*inch))
                
                datos_lista = datos['datos']
                if datos_lista:
                    # Seleccionar columnas principales (excluir objetos complejos)
                    primera_fila = datos_lista[0]
                    headers_principales = []
                    for key in primera_fila.keys():
                        valor = primera_fila[key]
                        # Incluir solo valores simples o strings cortos
                        if not isinstance(valor, (dict, list)) or (isinstance(valor, str) and len(valor) < 100):
                            headers_principales.append(key)
                    
                    if headers_principales:
                        # Detectar si es reporte de productos
                        es_reporte_productos = 'precio' in headers_principales and ('stock' in headers_principales or 'categoria' in headers_principales or 'marca' in headers_principales)
                        
                        # Si es reporte de productos, usar solo las columnas específicas
                        if es_reporte_productos:
                            columnas_productos = [
                                'nombre',
                                'precio',
                                'categoria',
                                'stock',
                                'marca',
                                'monto_total_vendido',
                                'veces_vendido'
                            ]
                            columnas_principales = [col for col in columnas_productos if col in headers_principales]
                        else:
                            # Filtrar columnas para mostrar solo las principales (excluir innecesarias)
                            columnas_principales = []
                            excluir_columnas = [
                                'id', 'id_venta', 'id_reporte',  # IDs no necesarios
                                'fecha_iso', 'fecha_formateada',  # Usar solo 'fecha'
                                'total_numero', 'monto_total',  # Usar versiones formateadas
                                'cliente', 'productos',  # Objetos complejos
                                'productos_count', 'total_productos_cantidad',  # Columnas eliminadas
                                'notas', 'direccion_entrega',  # Solo en detalles
                                'fecha_primera_compra', 'veces_comprado',  # Información extra
                                'precio_unitario_promedio',  # Usar precio_unitario formateado
                                'cantidad_vendida',  # Excluir cantidad_vendida, solo mostrar monto_total_vendido y veces_vendido
                            ]
                            
                            # Verificar si el usuario es administrador
                            es_admin = reporte.id_usuario and reporte.id_usuario.id_rol and reporte.id_usuario.id_rol.nombre.lower() == 'administrador'
                            
                            # Si no es admin, excluir columnas administrativas
                            if not es_admin:
                                excluir_columnas.extend(['estado', 'metodo_pago', 'cliente_email', 'cliente_telefono', 'cliente_direccion'])
                            
                            # Prioridades para columnas principales - NOMBRE SIEMPRE PRIMERO cuando hay productos
                            if 'nombre' in headers_principales:
                                prioridades = ['nombre', 'fecha', 'total', 'precio_total', 'precio_unitario', 'cantidad', 'categoria', 'cliente_nombre']
                            else:
                                prioridades = ['nombre', 'fecha', 'total', 'precio_total', 'precio_unitario', 'cantidad', 'categoria', 'cliente_nombre']
                            
                            # Si es admin, agregar columnas administrativas
                            if es_admin:
                                prioridades.extend(['estado', 'metodo_pago', 'cliente_email', 'cliente_telefono'])
                            
                            # SIEMPRE incluir 'nombre' primero si existe
                            if 'nombre' in headers_principales and 'nombre' not in excluir_columnas:
                                columnas_principales.append('nombre')
                            
                            # Luego agregar otras prioridades
                            for h in prioridades:
                                if h != 'nombre' and h in headers_principales and h not in excluir_columnas and h not in columnas_principales:
                                    columnas_principales.append(h)
                            
                            # Luego agregar otras columnas formateadas
                            for h in headers_principales:
                                if h not in columnas_principales and h not in excluir_columnas:
                                    # Priorizar versiones formateadas
                                    if any(h.endswith(sufijo) for sufijo in ['_formateado', '_display', '_nombre']):
                                        base = h.replace('_formateado', '').replace('_display', '').replace('_nombre', '')
                                        if base not in [c.replace('_formateado', '').replace('_display', '').replace('_nombre', '') for c in columnas_principales]:
                                            columnas_principales.append(h)
                                    elif not any(h.startswith(ex) or h == ex for ex in excluir_columnas):
                                        if not any(c.replace('_formateado', '').replace('_display', '').replace('_nombre', '') == h for c in columnas_principales):
                                            columnas_principales.append(h)
                            
                            # Si no hay suficientes columnas, agregar algunas básicas
                            if len(columnas_principales) < 2:
                                for h in ['fecha', 'total', 'nombre']:
                                    if h in headers_principales and h not in columnas_principales:
                                        columnas_principales.insert(0, h)
                        
                        tabla_data = [[h.replace('_', ' ').replace('formateado', '').replace('display', '').replace('_iso', '').replace('_numero', '').title().strip() for h in columnas_principales]]
                        
                        # Filas (limitar a 100 para PDF)
                        for item in datos_lista[:100]:
                            fila = []
                            for h in columnas_principales:
                                valor = item.get(h, '')
                                # Formatear valores
                                if isinstance(valor, (int, float)):
                                    if 'monto_total_vendido' in h.lower() or ('total' in h.lower() and 'vendido' in h.lower()):
                                        valor_str = f"Bs. {valor:,.2f}"
                                    elif 'precio' in h.lower():
                                        valor_str = f"Bs. {valor:,.2f}"
                                    elif 'veces_vendido' in h.lower() or 'cantidad_vendida' in h.lower() or 'stock' in h.lower():
                                        valor_str = f"{int(valor):,}"
                                    else:
                                        valor_str = f"{valor:,}"
                                elif isinstance(valor, bool):
                                    valor_str = 'Sí' if valor else 'No'
                                elif isinstance(valor, (dict, list)):
                                    valor_str = f"{len(valor)} items" if isinstance(valor, list) else "Ver detalles"
                                else:
                                    # Para nombres, mostrar completo pero limitar si es muy largo
                                    if 'nombre' in h.lower():
                                        valor_str = str(valor)  # Nombre completo
                                    else:
                                        # Limitar longitud para otras columnas
                                        valor_str = str(valor)[:60] + ('...' if len(str(valor)) > 60 else '')
                                fila.append(valor_str)
                            tabla_data.append(fila)
                        
                        # Crear tabla con ancho dinámico y mejor formato
                        num_cols = len(columnas_principales)
                        ancho_disponible = 7*inch  # Ancho disponible en A4 con márgenes
                        
                        # Anchos más específicos para columnas comunes - ajustados para mejor distribución
                        anchos_cols = []
                        for h in columnas_principales:
                            if 'nombre' in h.lower():
                                anchos_cols.append(1.8*inch)  # Nombre más ancho sin descripción
                            elif 'fecha' in h.lower():
                                anchos_cols.append(1.1*inch)
                            elif 'monto_total_vendido' in h.lower():
                                anchos_cols.append(1.4*inch)  # Monto total vendido más ancho
                            elif 'veces_vendido' in h.lower():
                                anchos_cols.append(1.1*inch)  # Veces vendido
                            elif 'total' in h.lower() or 'monto' in h.lower() or 'precio_total' in h.lower():
                                anchos_cols.append(1.1*inch)
                            elif 'precio_unitario' in h.lower() or 'precio' in h.lower():
                                anchos_cols.append(1.1*inch)
                            elif 'categoria' in h.lower():
                                anchos_cols.append(1.2*inch)
                            elif 'stock' in h.lower():
                                anchos_cols.append(0.9*inch)  # Stock
                            elif 'marca' in h.lower():
                                anchos_cols.append(1.2*inch)  # Marca
                            elif 'cantidad' in h.lower():
                                anchos_cols.append(1.0*inch)
                            elif 'cliente' in h.lower():
                                anchos_cols.append(1.8*inch)
                            else:
                                anchos_cols.append(1.0*inch)  # Ancho por defecto
                        
                        # Ajustar si hay muchas columnas
                        total_width = sum(anchos_cols)
                        if total_width > ancho_disponible:
                            factor = ancho_disponible / total_width
                            anchos_cols = [w * factor for w in anchos_cols]
                        
                        tabla = Table(tabla_data, colWidths=anchos_cols, repeatRows=1)
                        # Determinar alineación por columna
                        alignments = []
                        for h in columnas_principales:
                            if 'nombre' in h.lower():
                                alignments.append('LEFT')
                            elif 'precio' in h.lower() or 'monto' in h.lower() or 'total' in h.lower() or 'stock' in h.lower() or 'veces_vendido' in h.lower() or 'cantidad_vendida' in h.lower():
                                alignments.append('CENTER')
                            elif 'categoria' in h.lower() or 'marca' in h.lower():
                                alignments.append('CENTER')
                            else:
                                alignments.append('CENTER')
                        
                        # Crear estilos base
                        styles_list = [
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066FF')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Encabezados centrados
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Verticalmente centrado
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),  # Tamaño de fuente normal para headers
                            ('FONTSIZE', (0, 1), (-1, -1), 9),  # Tamaño de fuente normal para datos
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Más espacio vertical para headers
                            ('TOPPADDING', (0, 0), (-1, 0), 12),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),  # Más espacio vertical para datos
                            ('TOPPADDING', (0, 1), (-1, -1), 10),
                            ('LEFTPADDING', (0, 0), (-1, -1), 8),  # Más espacio lateral
                            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                            ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir salto de línea
                        ]
                        
                        # Aplicar alineación específica por columna
                        for idx, align in enumerate(alignments):
                            styles_list.append(('ALIGN', (idx, 1), (idx, -1), align))
                        
                        tabla.setStyle(TableStyle(styles_list))
                        story.append(tabla)
                        
                        if len(datos_lista) > 100:
                            story.append(Spacer(1, 0.2*inch))
                            story.append(Paragraph(f"<i>Nota: Se muestran 100 de {len(datos_lista)} registros totales.</i>", styles['Normal']))
        
        try:
            doc.build(story)
        except Exception as e:
            logger.error(f"Error al construir PDF: {str(e)}", exc_info=True)
            raise
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="reporte_{reporte.id_reporte}.pdf"'
        return response
    
    def _generar_excel(self, reporte: Reporte):
        """Generar Excel del reporte con formato mejorado"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Título (sin merge para evitar errores)
        ws['A1'] = reporte.nombre
        ws['A1'].font = Font(bold=True, size=16, color="0066FF")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Información del reporte (solo información relevante)
        # Convertir a zona horaria local si está en UTC
        fecha_generacion = reporte.fecha_generacion
        if timezone.is_aware(fecha_generacion):
            fecha_generacion = timezone.localtime(fecha_generacion)
        fecha_formateada = fecha_generacion.strftime('%d/%m/%Y %H:%M:%S')
        
        row = 3
        info_items = [
            ('Fecha de Generación:', fecha_formateada),
            ('Origen:', reporte.get_origen_comando_display()),

        ]
        
        if reporte.prompt:
            info_items.append(('Solicitud Original:', reporte.prompt[:200]))
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Datos del reporte
        datos = reporte.datos
        if isinstance(datos, dict):
            # Mostrar resumen si existe
            if 'resumen' in datos and datos['resumen']:
                ws[f'A{row}'] = 'RESUMEN'
                ws[f'A{row}'].font = Font(bold=True, size=12, color="0066FF")
                row += 1
                
                resumen = datos['resumen']
                # Campos a excluir del resumen
                campos_excluir = [
                    'compras_mostradas', 'compras_totales', 'categorias_analizadas', 
                    'clientes_mostrados', 'mensaje'
                ]
                
                # Campos prioritarios a mostrar primero
                campos_prioritarios = [
                    'total_general', 'total_general_formateado',
                    'cantidad_compras', 'promedio_compra', 'promedio_compra_formateado',
                    'max_compra', 'max_compra_formateado',
                    'min_compra', 'min_compra_formateado'
                ]
                
                # Primero agregar campos prioritarios formateados
                for key in campos_prioritarios:
                    if key in resumen and key not in campos_excluir:
                        value = resumen[key]
                        if isinstance(value, (dict, list)):
                            continue
                        
                        # Si es formateado, usarlo directamente
                        if key.endswith('_formateado'):
                            label = key.replace('_formateado', '').replace('_', ' ').title().strip()
                            ws[f'A{row}'] = label
                            ws[f'A{row}'].font = Font(bold=True)
                            ws[f'B{row}'] = str(value)
                            row += 1
                        elif isinstance(value, (int, float)):
                            # Formatear si no hay versión formateada
                            key_formateado = key + '_formateado'
                            if key_formateado not in resumen:
                                label = key.replace('_', ' ').title().strip()
                                ws[f'A{row}'] = label
                                ws[f'A{row}'].font = Font(bold=True)
                                if 'total' in key.lower() or 'monto' in key.lower() or 'precio' in key.lower() or 'compra' in key.lower():
                                    ws[f'B{row}'] = f"Bs. {value:,.2f}"
                                else:
                                    ws[f'B{row}'] = f"{value:,}"
                                row += 1
                
                # Luego agregar otros campos relevantes
                for key, value in resumen.items():
                    if key in campos_excluir or key in campos_prioritarios:
                        continue
                    if isinstance(value, (dict, list)):
                        continue
                    
                    # Si hay versión formateada, usar esa
                    key_formateado = key + '_formateado'
                    if key_formateado in resumen:
                        label = key.replace('_', ' ').title().strip()
                        ws[f'A{row}'] = label
                        ws[f'A{row}'].font = Font(bold=True)
                        ws[f'B{row}'] = str(resumen[key_formateado])
                        row += 1
                    elif not key.endswith('_formateado') and not key.endswith('_display'):
                        # Solo agregar si no es un campo formateado
                        label = key.replace('_', ' ').title().strip()
                        ws[f'A{row}'] = label
                        ws[f'A{row}'].font = Font(bold=True)
                        if isinstance(value, (int, float)):
                            if 'total' in key.lower() or 'monto' in key.lower() or 'precio' in key.lower() or 'compra' in key.lower():
                                ws[f'B{row}'] = f"Bs. {value:,.2f}"
                            else:
                                ws[f'B{row}'] = f"{value:,}"
                        else:
                            ws[f'B{row}'] = str(value)
                        row += 1
                
                row += 1
            
            # Mostrar datos principales
            if 'datos' in datos and datos['datos']:
                ws[f'A{row}'] = 'DETALLES'
                ws[f'A{row}'].font = Font(bold=True, size=12, color="0066FF")
                row += 1
                
                datos_lista = datos['datos']
                if datos_lista:
                    # Seleccionar columnas principales
                    primera_fila = datos_lista[0]
                    headers_principales = []
                    for key in primera_fila.keys():
                        valor = primera_fila[key]
                        if not isinstance(valor, (dict, list)) or (isinstance(valor, str) and len(valor) < 100):
                            headers_principales.append(key)
                    
                    if headers_principales:
                        # Detectar si es reporte de productos
                        es_reporte_productos = 'precio' in headers_principales and ('stock' in headers_principales or 'categoria' in headers_principales or 'marca' in headers_principales)
                        
                        # Si es reporte de productos, usar solo las columnas específicas
                        if es_reporte_productos:
                            columnas_productos = [
                                'nombre',
                                'precio',
                                'categoria',
                                'stock',
                                'marca',
                                'monto_total_vendido',
                                'veces_vendido'
                            ]
                            columnas_principales = [col for col in columnas_productos if col in headers_principales]
                        else:
                            # Filtrar columnas para mostrar solo las principales
                            columnas_principales = []
                            excluir_columnas = [
                                'id', 'id_venta', 'id_reporte',  # IDs no necesarios
                                'fecha_iso', 'fecha_formateada',  # Usar solo 'fecha'
                                'total_numero', 'monto_total',  # Usar versiones formateadas
                                'cliente', 'productos',  # Objetos complejos
                                'productos_count', 'total_productos_cantidad',  # Columnas eliminadas
                                'notas', 'direccion_entrega',  # Solo en detalles
                                'fecha_primera_compra', 'veces_comprado',  # Información extra
                                'precio_unitario_promedio',  # Usar precio_unitario formateado
                                'cantidad_vendida',  # Excluir cantidad_vendida, solo mostrar monto_total_vendido y veces_vendido
                            ]
                            
                            # Verificar si el usuario es administrador
                            es_admin = reporte.id_usuario and reporte.id_usuario.id_rol and reporte.id_usuario.id_rol.nombre.lower() == 'administrador'
                            
                            # Si no es admin, excluir columnas administrativas
                            if not es_admin:
                                excluir_columnas.extend(['estado', 'metodo_pago', 'cliente_email', 'cliente_telefono', 'cliente_direccion'])
                            
                            # Prioridades para columnas principales
                            if 'nombre' in headers_principales:
                                prioridades = ['nombre', 'fecha', 'total', 'precio_total', 'precio_unitario', 'cantidad', 'categoria', 'cliente_nombre']
                            else:
                                prioridades = ['nombre', 'fecha', 'total', 'precio_total', 'precio_unitario', 'cantidad', 'categoria', 'cliente_nombre']
                            
                            # Si es admin, agregar columnas administrativas
                            if es_admin:
                                prioridades.extend(['estado', 'metodo_pago', 'cliente_email', 'cliente_telefono'])
                            
                            # SIEMPRE incluir 'nombre' primero si existe
                            if 'nombre' in headers_principales and 'nombre' not in excluir_columnas:
                                columnas_principales.append('nombre')
                            
                            # Luego agregar otras prioridades
                            for h in prioridades:
                                if h != 'nombre' and h in headers_principales and h not in excluir_columnas and h not in columnas_principales:
                                    columnas_principales.append(h)
                            
                            # Luego agregar otras columnas formateadas
                            for h in headers_principales:
                                if h not in columnas_principales and h not in excluir_columnas:
                                    # Priorizar versiones formateadas
                                    if any(h.endswith(sufijo) for sufijo in ['_formateado', '_display', '_nombre']):
                                        base = h.replace('_formateado', '').replace('_display', '').replace('_nombre', '')
                                        if base not in [c.replace('_formateado', '').replace('_display', '').replace('_nombre', '') for c in columnas_principales]:
                                            columnas_principales.append(h)
                                    elif not any(h.startswith(ex) or h == ex for ex in excluir_columnas):
                                        if not any(c.replace('_formateado', '').replace('_display', '').replace('_nombre', '') == h for c in columnas_principales):
                                            columnas_principales.append(h)
                            
                            # Si no hay suficientes columnas, agregar algunas básicas
                            if len(columnas_principales) < 3:
                                for h in ['fecha', 'total', 'nombre', 'cliente_nombre']:
                                    if h in headers_principales and h not in columnas_principales:
                                        columnas_principales.insert(0, h)
                        
                        # Encabezados
                        header_row = row
                        for col, header in enumerate(columnas_principales, 1):
                            cell = ws.cell(row=header_row, column=col)
                            cell.value = header.replace('_', ' ').replace('formateado', '').replace('display', '').replace('_iso', '').replace('_numero', '').title().strip()
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                        
                        # Filas de datos
                        for item_idx, item in enumerate(datos_lista[:1000], 0):  # Limitar a 1000
                            data_row = header_row + 1 + item_idx
                            for col, header in enumerate(columnas_principales, 1):
                                cell = ws.cell(row=data_row, column=col)
                                valor = item.get(header, '')
                                
                                # Formatear valores
                                if isinstance(valor, (int, float)):
                                    if 'monto_total_vendido' in header.lower() or ('total' in header.lower() and 'vendido' in header.lower()) or ('monto' in header.lower() and 'vendido' in header.lower()):
                                        cell.value = valor
                                        cell.number_format = '$#,##0.00'  # Formato de moneda para monto total vendido
                                    elif 'total' in header.lower() or 'monto' in header.lower() or 'precio' in header.lower() or 'compra' in header.lower():
                                        cell.value = valor
                                        cell.number_format = '$#,##0.00'
                                    elif 'veces_vendido' in header.lower() or 'cantidad_vendida' in header.lower() or 'stock' in header.lower():
                                        cell.value = valor
                                        cell.number_format = '#,##0'  # Formato numérico sin decimales
                                    else:
                                        cell.value = valor
                                        cell.number_format = '#,##0'
                                elif isinstance(valor, bool):
                                    cell.value = 'Sí' if valor else 'No'
                                elif isinstance(valor, (dict, list)):
                                    cell.value = f"{len(valor)} items" if isinstance(valor, list) else "Ver detalles"
                                else:
                                    # Para nombres, mostrar completo sin truncar
                                    if 'nombre' in header.lower():
                                        cell.value = str(valor)  # Nombre completo
                                    else:
                                        cell.value = str(valor)[:100]  # Limitar longitud para otras columnas
                                
                                # Determinar alineación según el tipo de columna
                                if 'nombre' in header.lower():
                                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                                elif 'precio' in header.lower() or 'monto' in header.lower() or 'total' in header.lower() or 'stock' in header.lower() or 'veces_vendido' in header.lower() or 'cantidad_vendida' in header.lower():
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                elif 'categoria' in header.lower() or 'marca' in header.lower():
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                
                                cell.border = Border(
                                    left=Side(style='thin', color='CCCCCC'),
                                    right=Side(style='thin', color='CCCCCC'),
                                    top=Side(style='thin', color='CCCCCC'),
                                    bottom=Side(style='thin', color='CCCCCC')
                                )
                                
                                # Alternar colores de fila
                                if item_idx % 2 == 0:
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                                else:
                                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        # Ajustar ancho de columnas (evitar merged cells)
        from openpyxl.utils import get_column_letter
        
        # Obtener el número máximo de columnas usadas
        max_col = ws.max_column
        
        # Buscar la fila de encabezados (buscar desde arriba)
        header_row_num = None
        for row_idx in range(1, min(10, ws.max_row + 1)):  # Buscar en las primeras 10 filas
            cell = ws.cell(row=row_idx, column=1)
            if cell.value and ('DETALLES' in str(cell.value).upper() or 'RESUMEN' in str(cell.value).upper()):
                header_row_num = row_idx + 1  # La siguiente fila es la de encabezados
                break
        
        # Anchos específicos para columnas de productos
        anchos_columnas_productos = {
            'nombre': 30,
            'precio': 14,
            'categoria': 16,
            'stock': 10,
            'marca': 14,
            'monto_total_vendido': 20,
            'veces_vendido': 14
        }
        
        # Si no encontramos, buscar la fila con más celdas con texto (probablemente encabezados)
        if not header_row_num:
            for row_idx in range(1, min(20, ws.max_row + 1)):
                row_cells = [ws.cell(row=row_idx, column=col) for col in range(1, min(max_col + 1, 10))]
                if sum(1 for c in row_cells if c.value and isinstance(c.value, str) and len(str(c.value)) > 0) >= 3:
                    header_row_num = row_idx
                    break
        
        if not header_row_num:
            header_row_num = 1
        
        for col_num in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            # Obtener el header de esta columna para determinar ancho especial
            header_cell = ws.cell(row=header_row_num, column=col_num)
            header_text = str(header_cell.value or '').lower()
            
            # Verificar si la columna tiene celdas merged
            try:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            # Saltar merged cells
                            if hasattr(cell, 'value') and cell.value is not None:
                                cell_value = str(cell.value)
                                if len(cell_value) > max_length:
                                    max_length = len(cell_value)
                        except:
                            pass
            except:
                pass
            
            # Ajustar ancho según el tipo de columna
            # Primero verificar si es una columna de productos con ancho específico
            header_clean = header_text.replace('_', ' ').replace('formateado', '').replace('display', '').strip()
            if header_clean in anchos_columnas_productos:
                adjusted_width = anchos_columnas_productos[header_clean]
            elif 'nombre' in header_text:
                adjusted_width = min(max(max_length + 2, 25), 50)  # Nombre más ancho
            elif 'descripcion' in header_text or 'descripción' in header_text:
                adjusted_width = min(max(max_length + 2, 40), 60)  # Descripción más ancha
            elif 'monto_total_vendido' in header_text:
                adjusted_width = min(max(max_length + 2, 18), 25)  # Monto total vendido
            elif 'veces_vendido' in header_text:
                adjusted_width = min(max(max_length + 2, 12), 15)  # Veces vendido
            elif 'fecha' in header_text:
                adjusted_width = min(max(max_length + 2, 15), 20)
            elif 'total' in header_text or 'precio' in header_text or 'monto' in header_text:
                adjusted_width = min(max(max_length + 2, 12), 18)
            elif 'categoria' in header_text:
                adjusted_width = min(max(max_length + 2, 15), 25)
            elif 'marca' in header_text:
                adjusted_width = min(max(max_length + 2, 12), 20)
            elif 'stock' in header_text:
                adjusted_width = min(max(max_length + 2, 8), 12)  # Stock más compacto
            else:
                adjusted_width = min(max(max_length + 2, 12), 30)
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        try:
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_{reporte.id_reporte}.xlsx"'
            return response
        except Exception as e:
            logger.error(f"Error al generar Excel: {str(e)}", exc_info=True)
            raise


@method_decorator(csrf_exempt, name='dispatch')
class FiltrosInteligentesView(View):
    """
    CU19: Aplicar Filtros Inteligentes
    Usa IA para sugerir filtros relevantes
    """
    
    def post(self, request):
        """Obtener filtros inteligentes sugeridos"""
        try:
            data = json.loads(request.body)
            tipo_reporte = data.get('tipo_reporte', 'ventas')
            
            # Generar sugerencias usando análisis básico
            sugerencias = self._generar_sugerencias(tipo_reporte)
            
            return JsonResponse({
                'success': True,
                'sugerencias': sugerencias
            }, status=200)
            
        except Exception as e:
            logger.error(f"Error en FiltrosInteligentesView.post: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error interno: {str(e)}'
            }, status=500)
    
    def _generar_sugerencias(self, tipo: str) -> list:
        """Generar sugerencias de filtros"""
        sugerencias = []
        
        if tipo == 'ventas':
            # Analizar tendencias
            hoy = timezone.now().date()
            semana_pasada = hoy - timedelta(days=7)
            
            ventas_semana = Venta.objects.filter(
                fecha_venta__gte=semana_pasada,
                estado='completada'
            ).count()
            
            if ventas_semana > 10:
                sugerencias.append({
                    'filtro': 'fecha',
                    'tipo': 'periodo',
                    'valor': 'última semana',
                    'razon': f'Se registraron {ventas_semana} ventas en la última semana'
                })
            
            # Categorías más vendidas
            from django.db.models import Sum
            categorias_vendidas = DetalleVenta.objects.filter(
                venta__fecha_venta__gte=semana_pasada,
                producto__categoria__isnull=False
            ).values('producto__categoria__nombre').annotate(
                total=Sum('subtotal')
            ).order_by('-total')[:3]
            
            for cat in categorias_vendidas:
                sugerencias.append({
                    'filtro': 'categoria',
                    'tipo': 'valor',
                    'valor': cat['producto__categoria__nombre'],
                    'razon': f'Categoría con mayor movimiento: Bs. {cat["total"]:.2f}'
                })
        
        return sugerencias


@method_decorator(csrf_exempt, name='dispatch')
class OpcionesFiltrosView(View):
    """
    Vista auxiliar para obtener opciones de filtros para reportes
    Retorna listas de clientes, categorías, productos, etc. para usar en filtros
    """
    
    def get(self, request):
        """Obtener opciones para filtros de reportes"""
        try:
            # Verificar autenticación
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión'
                }, status=401)
            
            user_id = request.session.get('user_id')
            try:
                usuario = Usuario.objects.get(id=user_id)
                is_admin = usuario.id_rol and usuario.id_rol.nombre.lower() == 'administrador'
            except Usuario.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Usuario no encontrado'
                }, status=404)
            
            response_data = {
                'success': True,
                'is_admin': is_admin
            }
            
            # Solo admin puede ver lista de clientes
            if is_admin:
                clientes = Cliente.objects.select_related('id').all().order_by('id__nombre')
                clientes_list = [
                    {
                        'id': cliente.id.id,
                        'nombre': f"{cliente.id.nombre} {cliente.id.apellido or ''}".strip(),
                        'email': cliente.id.email
                    }
                    for cliente in clientes
                ]
                response_data['clientes'] = clientes_list
            
            # Obtener lista de categorías para filtro
            categorias = Categoria.objects.all().order_by('nombre')
            categorias_list = [
                {
                    'id': categoria.id_categoria,
                    'nombre': categoria.nombre
                }
                for categoria in categorias
            ]
            response_data['categorias'] = categorias_list
            
            # Métodos de pago disponibles
            response_data['metodos_pago'] = [
                {'value': 'stripe', 'label': 'Pago con Tarjeta (Stripe)'}
            ]
            
            # Estados disponibles
            response_data['estados'] = [
                {'value': 'pendiente', 'label': 'Pendiente'},
                {'value': 'completada', 'label': 'Completada'},
                {'value': 'cancelada', 'label': 'Cancelada'}
            ]
            
            return JsonResponse(response_data, status=200)
            
        except Exception as e:
            logger.error(f"Error en OpcionesFiltrosView.get: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error interno: {str(e)}'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class ListarReportesView(View):
    """Listar reportes generados"""
    
    def get(self, request):
        """Listar reportes del usuario"""
        try:
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión'
                }, status=401)
            
            user_id = request.session.get('user_id')
            if not user_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Usuario no autenticado'
                }, status=401)
            
            try:
                usuario = Usuario.objects.get(id=user_id)
            except Usuario.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Usuario no encontrado'
                }, status=404)
            
            # Obtener reportes del usuario
            reportes = Reporte.objects.filter(id_usuario=usuario).order_by('-fecha_generacion')[:50]
            
            datos = []
            for reporte in reportes:
                datos.append({
                    'id': reporte.id_reporte,
                    'nombre': reporte.nombre,
                    'tipo': reporte.tipo,
                    'formato': reporte.formato,
                    'fecha': reporte.fecha_generacion.isoformat(),
                    'origen': reporte.get_origen_comando_display(),
                    'pdf_url': f'/api/reportes/{reporte.id_reporte}/descargar/?formato=pdf' if reporte.formato == 'pdf' else None,
                    'excel_url': f'/api/reportes/{reporte.id_reporte}/descargar/?formato=excel'
                })
            
            return JsonResponse({
                'success': True,
                'reportes': datos
            }, status=200)
            
        except Exception as e:
            logger.error(f"Error en ListarReportesView.get: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error interno: {str(e)}'
            }, status=500)
